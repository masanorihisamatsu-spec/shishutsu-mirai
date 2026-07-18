"""
Excel家計簿（.xlsx）を解析する。

想定レイアウト: アクティブシートの1行目がヘッダー、2行目以降が明細。
必須列: 日付, 店舗名, 金額
任意列: カテゴリ（省略時は「その他」）, 支払方法（省略時は「その他」）

注意: 列名・シート構成は利用者の家計簿ごとに異なるため、
実データに合わせてヘッダー名の対応関係を調整すること。
"""

from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.services.imports.base import ParsedTransaction, ParseResult

REQUIRED_HEADERS = ("日付", "店舗名", "金額")
OPTIONAL_HEADERS = ("カテゴリ", "支払方法")


def parse(content: bytes) -> ParseResult:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001 - 壊れたExcelファイル等を一律ドメイン例外相当として扱う
        return ParseResult(transactions=[], errors=[f"Excelファイルを読み込めませんでした: {exc}"])

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ParseResult(transactions=[], errors=["データが見つかりません"])

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

    try:
        date_idx = header.index("日付")
        store_idx = header.index("店舗名")
        amount_idx = header.index("金額")
    except ValueError as exc:
        return ParseResult(transactions=[], errors=[f"必須列が見つかりません: {exc}"])

    category_idx = header.index("カテゴリ") if "カテゴリ" in header else None
    payment_idx = header.index("支払方法") if "支払方法" in header else None

    transactions: list[ParsedTransaction] = []
    errors: list[str] = []

    for row_number, row in enumerate(rows[1:], start=2):
        if row is None or all(cell is None for cell in row):
            continue
        try:
            category = row[category_idx] if category_idx is not None else None
            payment_method = row[payment_idx] if payment_idx is not None else None
            transactions.append(
                ParsedTransaction(
                    date=_parse_date(row[date_idx]),
                    store_name=str(row[store_idx]).strip(),
                    amount=_parse_amount(row[amount_idx]),
                    category=str(category).strip() if category else "その他",
                    payment_method=str(payment_method).strip() if payment_method else "その他",
                )
            )
        except (TypeError, ValueError, IndexError) as exc:
            errors.append(f"{row_number}行目: {exc}")

    return ParseResult(transactions=transactions, errors=errors)


def _parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return datetime.strptime(value.strip().replace("/", "-"), "%Y-%m-%d").date()
    raise ValueError(f"日付として解釈できません: {value!r}")


def _parse_amount(value: object) -> int:
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        digits = value.strip().replace(",", "").replace("円", "")
        return int(digits)
    raise ValueError(f"金額として解釈できません: {value!r}")
